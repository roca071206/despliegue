"""
UTILIDADES PARA EXPORTACION DE REPORTES
Modulo con funciones para exportar datos a PDF y Excel
"""

from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime


# ====== EXPORTACION A PDF ======
def exportar_pdf(titulo, columnas, datos, nombre_archivo):
    """
    Exporta datos a PDF usando WeasyPrint.

    ─── CORRECCIÓN ──────────────────────────────────────────────────────────
    La versión anterior usaba {{ 'now'|date:'...' }} en el template, que no
    funciona en Django — el filtro date solo opera sobre objetos de fecha,
    no sobre la cadena literal 'now'. La fecha nunca aparecía.
    Ahora se pasa fecha_generacion en el contexto desde Python.
    ─────────────────────────────────────────────────────────────────────────
    """
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    contexto = {
        'titulo':            titulo,
        'columnas':          columnas,
        'datos':             datos,
        'fecha_generacion':  fecha_generacion,   # ← corregido
    }

    html_string = render_to_string('Reportes/reporte_pdf.html', contexto)
    html_object = HTML(string=html_string, base_url='.')
    pdf_bytes   = html_object.write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response


# ====== EXPORTACION A EXCEL ======
def exportar_excel(titulo, columnas, datos, nombre_archivo):
    """
    Exporta datos a Excel usando openpyxl.

    ─── CORRECCIÓN ──────────────────────────────────────────────────────────
    La versión anterior calculaba el ancho de columna solo mirando el header
    y los datos, pero usaba chr(64 + col_num) para las letras — eso falla
    para más de 26 columnas (Z+1 = '[', no 'AA'). Se reemplazó por
    get_column_letter() de openpyxl que maneja correctamente columnas múltiples.

    Además se agrega fila de fecha de generación y ancho mínimo de 12.
    ─────────────────────────────────────────────────────────────────────────
    """
    from openpyxl.utils import get_column_letter

    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    num_cols = len(columnas)

    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"

    # ── Estilos ──
    COLOR_AZUL_OSCURO = '1E3A5F'
    COLOR_AZUL_MEDIO  = '4472C4'
    COLOR_AZUL_CLARO  = 'DCE6F1'
    COLOR_BLANCO      = 'FFFFFF'

    titulo_font     = Font(name='Arial', size=14, bold=True, color=COLOR_BLANCO)
    titulo_fill     = PatternFill(start_color=COLOR_AZUL_OSCURO, end_color=COLOR_AZUL_OSCURO, fill_type='solid')
    titulo_align    = Alignment(horizontal='center', vertical='center')

    fecha_font      = Font(name='Arial', size=9, italic=True, color='555555')
    fecha_align     = Alignment(horizontal='right', vertical='center')

    header_font     = Font(name='Arial', size=10, bold=True, color=COLOR_BLANCO)
    header_fill     = PatternFill(start_color=COLOR_AZUL_MEDIO, end_color=COLOR_AZUL_MEDIO, fill_type='solid')
    header_align    = Alignment(horizontal='center', vertical='center', wrap_text=False)

    data_align      = Alignment(horizontal='left', vertical='center')
    data_align_num  = Alignment(horizontal='right', vertical='center')
    data_fill_par   = PatternFill(start_color=COLOR_AZUL_CLARO, end_color=COLOR_AZUL_CLARO, fill_type='solid')
    data_border     = Border(
        left=Side(style='thin'),   right=Side(style='thin'),
        top=Side(style='thin'),    bottom=Side(style='thin'),
    )
    last_col_letter = get_column_letter(num_cols)

    # ── Fila 1: Título ──
    worksheet.merge_cells(f'A1:{last_col_letter}1')
    cell = worksheet['A1']
    cell.value     = titulo
    cell.font      = titulo_font
    cell.fill      = titulo_fill
    cell.alignment = titulo_align
    worksheet.row_dimensions[1].height = 28

    # ── Fila 2: Fecha de generación ──
    worksheet.merge_cells(f'A2:{last_col_letter}2')
    cell = worksheet['A2']
    cell.value     = f'Generado el: {fecha_generacion}'
    cell.font      = fecha_font
    cell.alignment = fecha_align
    worksheet.row_dimensions[2].height = 16

    # ── Fila 3: Encabezados ──
    for col_num, columna in enumerate(columnas, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value     = columna
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = data_border
    worksheet.row_dimensions[3].height = 22

    # ── Filas 4+: Datos ──
    for row_num, fila in enumerate(datos, 4):
        valores = list(fila.values()) if isinstance(fila, dict) else list(fila)
        for col_num, valor in enumerate(valores, 1):
            cell        = worksheet.cell(row=row_num, column=col_num)
            cell.value  = valor
            cell.border = data_border
            # Números alineados a la derecha
            if isinstance(valor, (int, float)):
                cell.alignment = data_align_num
            else:
                cell.alignment = data_align
            # Filas alternas
            if (row_num - 4) % 2 == 0:
                cell.fill = data_fill_par
        worksheet.row_dimensions[row_num].height = 16

    # ── Fila total al final ──
    total_row = len(datos) + 4
    worksheet.merge_cells(f'A{total_row}:{last_col_letter}{total_row}')
    cell_total           = worksheet[f'A{total_row}']
    cell_total.value     = f'Total de registros: {len(datos)}'
    cell_total.font      = Font(name='Arial', size=10, bold=True, color=COLOR_AZUL_OSCURO)
    cell_total.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[total_row].height = 18

    # ── Ancho automático de columnas ──
    for col_num, columna in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)
        max_len    = len(str(columna))

        for row in worksheet.iter_rows(min_row=4, max_row=len(datos) + 3,
                                       min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

        # Ancho mínimo 12, máximo 45
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    # ── Freeze pane en fila 4 (encabezados siempre visibles) ──
    worksheet.freeze_panes = 'A4'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    workbook.save(response)
    return response